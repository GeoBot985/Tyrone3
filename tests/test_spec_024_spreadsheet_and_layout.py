import os
import tempfile
import unittest
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from rag.ingest import ingest_document
from rag.layout_normalizer import normalize_layout_aware_text
from rag.spreadsheet_extractor import extract_xlsx_structured


class TestSpec024SpreadsheetAndLayout(unittest.TestCase):
    def _temp_xlsx(self) -> str:
        handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        handle.close()
        self.addCleanup(lambda: os.path.exists(handle.name) and os.remove(handle.name))
        return handle.name

    def test_xlsx_extractor_preserves_header_row_mapping(self):
        path = self._temp_xlsx()
        wb = Workbook()
        ws = wb.active
        ws.title = "Claims"
        ws.append(["Date", "Member", "Provider", "Amount"])
        ws.append(["02 January 2026", "George H (Jnr)", "Barnard Vd Merwe Theron", "R672,12"])
        wb.save(path)

        result = extract_xlsx_structured(path)

        self.assertTrue(result["success"])
        self.assertEqual(result["method"], "openpyxl_structured")
        self.assertEqual(result["sheet_count"], 1)
        self.assertEqual(result["row_count"], 1)
        self.assertIn("[Sheet: Claims | Range: A2:D2 | Row: 2 | Region: table_row]", result["text"])
        self.assertIn("Provider: Barnard Vd Merwe Theron", result["text"])
        self.assertEqual(result["row_records"][1]["cell_range"], "A2:D2")
        self.assertEqual(result["row_records"][1]["region_type"], "table_row")

    def test_xlsx_extractor_handles_blank_rows_before_header(self):
        path = self._temp_xlsx()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append([None, None, None])
        ws.append([None, None, None])
        ws.append(["Date", "Member", "Amount"])
        ws.append(["03 January 2026", "George R (Snr)", "R120,00"])
        wb.save(path)

        result = extract_xlsx_structured(path)

        self.assertTrue(result["success"])
        self.assertEqual(result["sheets"][0]["header_row_index"], 3)
        self.assertIn("Member: George R (Snr)", result["text"])

    def test_xlsx_extractor_warns_when_merged_cells_exist(self):
        path = self._temp_xlsx()
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged"
        ws.append(["Header", "Value"])
        ws.append(["A", "1"])
        ws.merge_cells("A2:B2")
        wb.save(path)

        result = extract_xlsx_structured(path)

        self.assertTrue(result["success"])
        self.assertTrue(any("Merged cells" in warning for warning in result["warnings"]))

    def test_layout_normalizer_reflows_two_column_text(self):
        sample = "\n".join([
            "Left heading          Right heading",
            "Left body one         Right body one",
            "Left body two         Right body two",
            "Left body three       Right body three",
        ])

        result = normalize_layout_aware_text(sample, "pdf")

        self.assertEqual(result["layout_mode"], "column_reflow")
        self.assertTrue(result["applied"])
        self.assertLess(result["text"].find("Left body three"), result["text"].find("Right heading"))

    def test_layout_normalizer_leaves_plain_text_as_plain(self):
        sample = "This is a normal paragraph.\nIt should stay in order.\nNo column tricks here."

        result = normalize_layout_aware_text(sample, "docx")

        self.assertEqual(result["layout_mode"], "plain")
        self.assertFalse(result["applied"])
        self.assertIn("It should stay in order.", result["text"])

    def test_layout_normalizer_preserves_table_like_rows(self):
        sample = "\n".join([
            "Date      Member      Amount",
            "2026-01-01  George     100",
            "2026-01-02  Cornelia   250",
        ])

        result = normalize_layout_aware_text(sample, "pdf")

        self.assertEqual(result["layout_mode"], "table_like")
        self.assertIn("Date | Member | Amount", result["text"])

    @patch("rag.ingest.insert_chunk")
    @patch("rag.ingest.insert_document")
    @patch("rag.ingest.embed_text", return_value=[0.1, 0.2, 0.3])
    def test_ingest_xlsx_uses_structured_extractor(self, _mock_embed, _mock_insert_document, _mock_insert_chunk):
        path = self._temp_xlsx()
        wb = Workbook()
        ws = wb.active
        ws.title = "Claims"
        ws.append(["Member", "Provider", "Description"])
        ws.append(["George H (Jnr)", "Barnard Vd Merwe Theron", "GP Consultation"])
        wb.save(path)

        result = ingest_document(path, MagicMock(), document_name="claims.xlsx")

        self.assertEqual(result["file_type"], "xlsx")
        self.assertEqual(result["ingestion_method"], "openpyxl_structured")
        self.assertEqual(result["provenance"]["primary_extractor"], "openpyxl_structured")
        self.assertEqual(result["provenance"]["sheet_count"], 1)
        self.assertEqual(result["provenance"]["row_count"], 1)
        self.assertTrue(result["provenance"]["has_table_rows"])
        self.assertEqual(result["provenance"]["region_counts"].get("table_row"), 1)
        self.assertGreater(result["chunk_count"], 0)

    def test_xlsx_summary_rows_are_classified_separately(self):
        path = self._temp_xlsx()
        wb = Workbook()
        ws = wb.active
        ws.title = "Claims"
        ws.append(["Member", "Provider", "Amount"])
        ws.append(["George H (Jnr)", "GP Consultation", "100"])
        ws.append(["Total", "", "100"])
        wb.save(path)

        result = extract_xlsx_structured(path)

        region_types = [row["region_type"] for row in result["row_records"]]
        self.assertIn("summary_block", region_types)

    def test_ingest_xls_fails_clearly(self):
        handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
        handle.write(b"legacy")
        handle.close()
        self.addCleanup(lambda: os.path.exists(handle.name) and os.remove(handle.name))

        with self.assertRaisesRegex(Exception, "Legacy .xls spreadsheets are not yet supported") as ctx:
            ingest_document(handle.name, MagicMock(), document_name="legacy.xls")

        self.assertEqual(ctx.exception.provenance["failure_stage"], "file_validation")
        self.assertEqual(ctx.exception.provenance["primary_extractor"], "legacy_xls_unsupported")
