from __future__ import annotations

import os
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from app.services import chat_orchestrator as co
from app.services import confidence as conf
from app.services import rag_service
from app.services import session_grounding as sg
from eval.common import temporary_eval_db
from rag import db as rag_db
from rag import search as rag_search
from rag.spreadsheet_extractor import (
    _classify_region,
    _detect_header_row,
    _find_used_bounds,
    _format_cell_value,
    _is_numeric_like,
    extract_xlsx_structured,
)

REAL_OS_PATH_EXISTS = os.path.exists


def _sample_document(document_id: str = "doc1", name: str = "Demo Doc") -> dict:
    return {
        "document_id": document_id,
        "document_name": name,
        "source_path": "/tmp/demo.txt",
        "file_hash": f"hash-{document_id}",
        "file_size_bytes": 10,
        "ingested_at": "2026-01-01T00:00:00",
        "chunk_count": 2,
    }


def _sample_chunk(
    chunk_id: str,
    document_id: str,
    index: int,
    text: str,
    embedding: list[float],
    region_type: str | None = None,
) -> dict:
    return {
        "chunk_id": chunk_id,
        "document_id": document_id,
        "chunk_index": index,
        "text": text,
        "embedding": embedding,
        "region_type": region_type,
    }


class _FakeConn:
    def close(self) -> None:
        return None


def test_confidence_branches():
    refusal = conf.build_refusal_confidence(
        coverage_mode="coverage_required", coverage_truncated=True, reason="custom_reason"
    )
    assert refusal["label"] == "low"
    assert refusal["reason_codes"] == ["custom_reason"]

    retrieval_error = conf.compute_document_confidence(
        chunks_used_for_prompt=[],
        retrieval_metrics=None,
        retrieval_error="boom",
        coverage_mode="narrow_lookup",
        coverage_truncated=False,
        skip_llm=True,
    )
    assert retrieval_error["score"] == 0.1
    assert retrieval_error["reason_codes"] == ["retrieval_error"]

    no_chunks_skip = conf.compute_document_confidence(
        chunks_used_for_prompt=[],
        retrieval_metrics=None,
        retrieval_error=None,
        coverage_mode="narrow_lookup",
        coverage_truncated=False,
        skip_llm=True,
    )
    assert no_chunks_skip["score"] == 0.12
    assert no_chunks_skip["label"] == "low"

    no_chunks_answered = conf.compute_document_confidence(
        chunks_used_for_prompt=[],
        retrieval_metrics=None,
        retrieval_error=None,
        coverage_mode="narrow_lookup",
        coverage_truncated=False,
        skip_llm=False,
    )
    assert no_chunks_answered["score"] == 0.18

    high = conf.compute_document_confidence(
        chunks_used_for_prompt=[
            {"score": 0.95, "lexical_score": 0.9},
        ],
        retrieval_metrics={},
        retrieval_error=None,
        coverage_mode="narrow_lookup",
        coverage_truncated=False,
        skip_llm=False,
    )
    assert "single_chunk_only" in high["reason_codes"]
    assert "strong_top_scores" in high["reason_codes"]

    low = conf.compute_document_confidence(
        chunks_used_for_prompt=[
            {"score": 0.2, "lexical_score": 0.1},
            {"score": 0.15, "lexical_score": 0.0},
        ],
        retrieval_metrics={"verification_status": "all_discarded", "bounded_negative_mode": True},
        retrieval_error=None,
        coverage_mode="coverage_required",
        coverage_truncated=True,
        skip_llm=False,
    )
    assert "multiple_verified_chunks" in low["reason_codes"]
    assert "weak_score_tail" in low["reason_codes"]
    assert "coverage_query_with_limited_evidence" in low["reason_codes"]
    assert "coverage_truncated" in low["reason_codes"]
    assert "verification_losses" in low["reason_codes"]
    assert "weak_lexical_support" in low["reason_codes"]


@pytest.mark.asyncio
async def test_session_grounding_build_usage_and_format(monkeypatch):
    monkeypatch.setattr(sg.tzlocal, "get_localzone_name", lambda: "Africa/Johannesburg")

    async def fake_get_models():
        return ["fake-model", "other-model"], ""

    monkeypatch.setattr(sg, "get_models", fake_get_models)

    context = await sg.build_session_grounding(active_mode="chat", selected_model="fake-model")
    assert context["model_available"] is True
    assert sg.get_session_grounding() == context

    sg.increment_session_usage(3, 4)
    updated = sg.get_session_grounding()
    assert updated["session_turn_count"] == 1
    assert updated["session_total_tokens_est"] == 7

    formatted = sg.format_grounding_for_debug(context)
    assert "Session Grounding" in formatted
    assert "model: fake-model" in formatted
    assert "(WARNING: Not available)" not in formatted


@pytest.mark.asyncio
async def test_session_grounding_timezone_fallback_and_warning(monkeypatch):
    monkeypatch.setattr(
        sg.tzlocal,
        "get_localzone_name",
        lambda: (_ for _ in ()).throw(RuntimeError("tzlocal failed")),
    )

    async def fake_get_models():
        return ["some-other-model"], ""

    monkeypatch.setattr(sg, "get_models", fake_get_models)

    context = await sg.build_session_grounding(active_mode="personal", selected_model="missing")
    assert context["timezone"] == "UTC"
    assert context["model_available"] is False
    assert "(WARNING: Not available)" in sg.format_grounding_for_debug(context)
    assert sg.format_grounding_for_debug({}) == "No grounding context initialized."


def test_rag_db_round_trip(tmp_path):
    db_path = tmp_path / "rag.db"
    conn = rag_db.get_connection(str(db_path))
    rag_db.init_db(conn)

    doc = _sample_document()
    rag_db.insert_document(conn, doc)
    rag_db.insert_chunk(
        conn,
        _sample_chunk("chunk-1", doc["document_id"], 0, "Alpha Beta", [0.1, 0.2, 0.3]),
    )
    rag_db.insert_chunk(
        conn,
        _sample_chunk("chunk-2", doc["document_id"], 1, "Gamma Delta", [0.4, 0.5, 0.6]),
    )

    listed = rag_db.list_documents(conn)
    assert listed[0]["document_id"] == doc["document_id"]
    assert rag_db.get_document_by_id(conn, doc["document_id"])["document_name"] == "Demo Doc"
    assert rag_db.get_document_by_hash(conn, doc["file_hash"])["document_id"] == doc["document_id"]
    assert rag_db.find_exact_chunk(conn, doc["document_id"], 0, "Alpha Beta")["chunk_id"] == "chunk-1"
    assert rag_db.find_exact_chunk(conn, doc["document_id"], 99, "Missing") is None

    embeddings = rag_db.get_all_embeddings(conn)
    assert len(embeddings) == 2
    assert embeddings[0]["embedding"][0] == pytest.approx(0.1)
    assert embeddings[0]["embedding"][1] == pytest.approx(0.2)
    assert embeddings[0]["embedding"][2] == pytest.approx(0.3)
    assert rag_db.get_all_chunks_for_document(conn, doc["document_id"])[0]["chunk_index"] == 0

    stats = rag_db.get_corpus_stats(conn)
    assert stats["total_documents"] == 1
    assert stats["total_chunks"] == 2

    assert rag_db.delete_document(conn, doc["document_id"]) is True
    assert rag_db.get_document_by_id(conn, doc["document_id"]) is None
    assert rag_db.get_corpus_stats(conn)["total_documents"] == 0

    rag_db.clear_corpus(conn)
    assert rag_db.get_corpus_stats(conn)["total_chunks"] == 0
    conn.close()


def test_spreadsheet_helpers_and_success_extract(tmp_path):
    assert _is_numeric_like(1) is True
    assert _is_numeric_like(1.5) is True
    assert _is_numeric_like(Decimal("2.5")) is True
    assert _is_numeric_like(True) is False
    assert _format_cell_value(None) == ""
    assert _format_cell_value(datetime(2026, 1, 1, 14, 30, 0)) == "01 January 2026 14:30:00"
    assert _format_cell_value(date(2026, 1, 1)) == "01 January 2026"
    assert _format_cell_value(time(9, 5, 0)) == "09:05:00"

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws["A1"] = "Name"
    ws["B1"] = "Amount"
    ws["A2"] = "Alice"
    ws["B2"] = 12.5
    ws["A3"] = "Bob"
    ws["B3"] = 20
    path = tmp_path / "expenses.xlsx"
    wb.save(path)

    loaded = Workbook()
    sheet = loaded.active
    sheet.title = "Demo"
    sheet["A1"] = "Name"
    sheet["B1"] = "Amount"
    sheet["A2"] = "Alice"
    sheet["B2"] = 10
    assert _find_used_bounds(sheet) == (1, 2, 1, 2)
    header_row, headers, warnings, used = _detect_header_row(sheet, 1, 2, 1, 2)
    assert header_row == 1
    assert headers == ["Name", "Amount"]
    assert warnings == []
    assert used is True
    assert _classify_region(headers, ["Alice", "10"], "Name: Alice | Amount: 10") == "pivot_like"
    assert _classify_region(headers, ["Alice", ""], "Name: Alice") == "table_row"
    assert _classify_region(headers, ["", ""], "") == "unknown"
    assert _classify_region(headers, ["Total", "10"], "sum total") == "summary_block"

    result = extract_xlsx_structured(str(path))
    assert result["success"] is True
    assert result["sheet_count"] == 1
    assert result["row_count"] == 2
    assert result["header_detection_used"] is True
    assert sum(result["region_counts"].values()) >= 2


def test_spreadsheet_extract_edge_cases(tmp_path, monkeypatch):
    class _HiddenSheet:
        sheet_state = "hidden"

    class _HiddenWorkbook:
        worksheets = [_HiddenSheet()]

    monkeypatch.setattr("rag.spreadsheet_extractor.os.path.exists", lambda *_args, **_kwargs: True)
    monkeypatch.setattr("rag.spreadsheet_extractor.load_workbook", lambda *_args, **_kwargs: _HiddenWorkbook())
    hidden_result = extract_xlsx_structured(str(tmp_path / "hidden.xlsx"))
    assert hidden_result["success"] is False
    assert "usable sheets" in hidden_result["error"]

    empty_wb = Workbook()
    empty_ws = empty_wb.active
    empty_ws.title = "Empty"
    empty_ws["A1"] = "Header"
    empty_path = tmp_path / "empty.xlsx"
    empty_wb.save(empty_path)
    monkeypatch.setattr("rag.spreadsheet_extractor.load_workbook", load_workbook)
    empty_wb = Workbook()
    empty_wb.active.title = "Empty"
    empty_wb.active["A1"].value = None
    empty_wb.active["B1"].value = None
    empty_wb.save(empty_path)
    empty_result = extract_xlsx_structured(str(empty_path))
    assert empty_result["success"] is False
    assert "no extractable rows" in empty_result["error"]

    monkeypatch.setattr("rag.spreadsheet_extractor.os.path.exists", REAL_OS_PATH_EXISTS)
    missing_result = extract_xlsx_structured(str(tmp_path / "missing.xlsx"))
    assert missing_result["success"] is False
    assert "File not found" in missing_result["error"]

    monkeypatch.setattr(
        "rag.spreadsheet_extractor.load_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("boom")),
    )
    failed_result = extract_xlsx_structured(str(empty_path))
    assert failed_result["success"] is False
    assert "Failed to read XLSX workbook" in failed_result["error"]


def test_search_helpers_and_results(tmp_path, monkeypatch):
    assert rag_search.detect_retrieval_mode("") == "default"
    assert rag_search.detect_retrieval_mode("list all entries") == "enumeration"
    assert rag_search._query_region_mode("top expenses") == "table_row_preferred"
    assert rag_search._query_region_mode("sum values") == "summary_allowed"
    assert rag_search._strong_lexical_hit(["exact_phrase"]) is True
    assert rag_search._strong_lexical_hit(["weak"]) is False
    assert rag_search.normalize_token("buses") == "bus"
    assert rag_search.normalize_token("cities") == "city"
    assert rag_search._extract_query_phrases('"quoted phrase" and more') == ["quoted phrase"]
    assert rag_search._extract_query_phrases("three word query") == ["three word query"]

    chunks = [
        {"document_id": "d1", "chunk_index": 0, "text": "Alpha", "document_name": "Doc A", "score": 0.8},
        {"document_id": "d1", "chunk_index": 0, "text": "Alpha", "document_name": "Doc A", "score": 0.8},
        {"document_id": "d2", "chunk_index": 1, "text": "Beta", "document_name": "Doc B", "score": 0.6},
    ]
    compacted = rag_service.compact_retrieved_chunks_for_prompt(chunks, "table")
    assert len(compacted) == 2
    selected, truncated, reason = rag_service._apply_coverage_stabilization(
        [
            {"score": 0.9},
            {"score": 0.7},
            {"score": 0.5},
            {"score": 0.2},
        ],
        max_chunks=3,
    )
    assert truncated is True
    assert reason in {"score_drop_stabilized", "max_coverage_chunks_reached", "low_score_floor_reached"}
    assert selected

    db_path = tmp_path / "search.db"
    conn = rag_db.get_connection(str(db_path))
    rag_db.init_db(conn)
    doc1 = _sample_document("doc1", "Expenses")
    doc2 = _sample_document("doc2", "Summary")
    rag_db.insert_document(conn, doc1)
    rag_db.insert_document(conn, doc2)
    rag_db.insert_chunk(
        conn,
        _sample_chunk("c1", "doc1", 0, "Alice 100", [1.0, 0.0], "table_row"),
    )
    rag_db.insert_chunk(
        conn,
        _sample_chunk("c2", "doc1", 1, "Total 100", [1.0, 0.0], "summary_block"),
    )
    rag_db.insert_chunk(
        conn,
        _sample_chunk("c3", "doc2", 0, "Bob 200", [1.0, 0.0], "pivot_like"),
    )

    monkeypatch.setattr(rag_search, "embed_text", lambda _query: [1.0, 0.0])

    enumeration = rag_search.search(
        conn,
        "list all entries",
        top_k=2,
        candidate_pool_size=5,
        per_doc_cap=2,
        retrieval_mode="enumeration",
        lexical_match_cap=5,
    )
    assert enumeration["metrics"]["retrieval_mode"] == "enumeration"
    assert enumeration["results"]

    summary = rag_search.search(
        conn,
        "sum values",
        top_k=2,
        candidate_pool_size=5,
        per_doc_cap=2,
        retrieval_mode="default",
        lexical_match_cap=5,
    )
    assert summary["metrics"]["region_mode"] == "summary_allowed"
    assert summary["results"][0]["text"] in {"Total 100", "Alice 100", "Bob 200"}

    table_pref = rag_search.search(
        conn,
        "top expenses",
        top_k=2,
        candidate_pool_size=5,
        per_doc_cap=2,
        retrieval_mode="default",
        lexical_match_cap=5,
    )
    assert table_pref["metrics"]["region_mode"] == "table_row_preferred"
    assert table_pref["results"][0]["region_type"] == "table_row"
    conn.close()


def test_rag_service_round_trip_and_missing_db(tmp_path, monkeypatch):
    missing = tmp_path / "missing.db"
    monkeypatch.setattr(rag_service, "DB_PATH", str(missing))
    assert rag_service.get_rag_context("hello")["error"] == "knowledge base empty / unavailable"
    assert rag_service.get_full_document_content("doc-x")["error"] == "knowledge base empty / unavailable"

    db_path = tmp_path / "service.db"
    with temporary_eval_db(db_path):
        conn = rag_db.get_connection(str(db_path))
        rag_db.init_db(conn)
        doc = _sample_document("doc1", "Document One")
        rag_db.insert_document(conn, doc)
        rag_db.insert_chunk(
            conn,
            _sample_chunk("chunk-1", "doc1", 0, "Alice 100", [1.0, 0.0], "table_row"),
        )
        rag_db.insert_chunk(
            conn,
            _sample_chunk("chunk-2", "doc1", 1, "Total 100", [1.0, 0.0], "summary_block"),
        )
        conn.close()

        monkeypatch.setattr(rag_search, "embed_text", lambda _query: [1.0, 0.0])
        context = rag_service.get_rag_context(
            "list all entries", top_k=2, document_ids=["doc1"], response_format="table"
        )
        assert context["error"] is None
        assert context["chunks"]
        assert context["chunks_for_prompt"]
        assert context["metrics"]["retrieval_chunks_used_for_prompt"] >= 1

        full = rag_service.get_full_document_content("doc1")
        assert full["document_name"] == "Document One"
        assert full["full_text"] == "Alice 100\nTotal 100"
        assert full["chunk_count"] == 2
        assert rag_service.get_full_document_content("missing")["error"] == "Document with ID missing not found."

        stats = rag_service.get_corpus_stats_service()
        assert stats["ok"] is True
        assert stats["stats"]["total_documents"] == 1

        listed = rag_service.list_indexed_documents()
        assert listed["ok"] is True
        assert listed["documents"][0]["document_id"] == "doc1"

        deleted = rag_service.delete_document_service("doc1")
        assert deleted["ok"] is True
        assert deleted["deleted"] is True

        cleared = rag_service.clear_corpus_service()
        assert cleared["ok"] is True
