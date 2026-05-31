from __future__ import annotations

from collections import Counter
from datetime import date, datetime, time
from decimal import Decimal
import os
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.config import SPREADSHEET_HEADER_SCAN_LIMIT


def _is_numeric_like(value) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def _format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d %B %Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d %B %Y")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def _find_used_bounds(sheet) -> tuple[int, int, int, int] | None:
    min_row = None
    max_row = 0
    min_col = None
    max_col = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if min_row is None or cell.row < min_row:
                min_row = cell.row
            if min_col is None or cell.column < min_col:
                min_col = cell.column
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
    if min_row is None or min_col is None:
        return None
    return min_row, max_row, min_col, max_col


def _row_values(sheet, row_index: int, min_col: int, max_col: int) -> list[str]:
    values = []
    for col_idx in range(min_col, max_col + 1):
        values.append(_format_cell_value(sheet.cell(row=row_index, column=col_idx).value))
    return values


def _detect_header_row(sheet, min_row: int, max_row: int, min_col: int, max_col: int) -> tuple[int | None, list[str], list[str], bool]:
    warnings: list[str] = []
    scan_to = min(max_row, min_row + SPREADSHEET_HEADER_SCAN_LIMIT - 1)
    for row_index in range(min_row, scan_to + 1):
        values = _row_values(sheet, row_index, min_col, max_col)
        non_empty = [value for value in values if value]
        if len(non_empty) < 2:
            continue
        non_numeric = [value for value in non_empty if not value.replace(",", "").replace(".", "", 1).isdigit()]
        if len(non_numeric) >= max(2, len(non_empty) // 2):
            headers = []
            for offset, value in enumerate(values, start=min_col):
                headers.append(value or f"Column_{get_column_letter(offset)}")
            return row_index, headers, warnings, True

    generated_headers = [f"Column_{get_column_letter(col_idx)}" for col_idx in range(min_col, max_col + 1)]
    warnings.append("No plausible header row found; generated column headers were used.")
    return None, generated_headers, warnings, False


def _classify_region(headers: list[str], values: list[str], row_text: str) -> str:
    lowered = row_text.lower()
    if any(keyword in lowered for keyword in ("sum", "total", "average", "avg")):
        return "summary_block"

    non_empty_values = [value for value in values if value]
    if len(non_empty_values) <= max(2, len(headers) // 2):
        has_numeric = any(value.replace(",", "").replace(".", "", 1).isdigit() for value in non_empty_values)
        has_alpha = any(re.search(r"[a-zA-Z]", value) for value in non_empty_values)
        if has_numeric and has_alpha:
            return "pivot_like"

    if non_empty_values:
        return "table_row"

    return "unknown"


def extract_xlsx_structured(path: str) -> dict:
    file_type = "xlsx"
    if not os.path.exists(path):
        return {
            "success": False,
            "text": "",
            "file_type": file_type,
            "method": "openpyxl_structured",
            "error": f"File not found: {path}",
        }

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        return {
            "success": False,
            "text": "",
            "file_type": file_type,
            "method": "openpyxl_structured",
            "error": f"Failed to read XLSX workbook: {exc}",
        }

    visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    if not visible_sheets:
        return {
            "success": False,
            "text": "",
            "file_type": file_type,
            "method": "openpyxl_structured",
            "error": "XLSX workbook could not identify any usable sheets.",
            "sheet_count": 0,
            "row_count": 0,
            "column_count_by_sheet": {},
            "sheets": [],
            "sheet_names": [],
            "skipped_objects": [],
            "warnings": [],
            "header_detection_used": False,
            "header_detection_warnings": [],
            "row_records": [],
            "region_counts": {},
        }

    text_blocks: list[str] = []
    row_records: list[dict] = []
    sheets_summary: list[dict] = []
    warnings: list[str] = []
    header_detection_warnings: list[str] = []
    column_count_by_sheet: dict[str, int] = {}
    sheet_names: list[str] = []
    total_row_count = 0
    merged_present = any(bool(sheet.merged_cells.ranges) for sheet in visible_sheets)

    for sheet in visible_sheets:
        bounds = _find_used_bounds(sheet)
        if not bounds:
            continue

        min_row, max_row, min_col, max_col = bounds
        header_row_index, headers, header_warnings, _header_detection_used = _detect_header_row(
            sheet,
            min_row,
            max_row,
            min_col,
            max_col,
        )
        header_detection_warnings.extend(header_warnings)
        column_count_by_sheet[sheet.title] = max_col - min_col + 1
        sheet_names.append(sheet.title)

        region_counts = Counter()
        data_start_row = (header_row_index + 1) if header_row_index is not None else min_row
        data_row_count = 0

        if header_row_index is not None:
            header_values = _row_values(sheet, header_row_index, min_col, max_col)
            header_range = f"{get_column_letter(min_col)}{header_row_index}:{get_column_letter(max_col)}{header_row_index}"
            header_pairs = [f"{header}: {value}" for header, value in zip(headers, header_values) if value]
            row_records.append({
                "sheet_name": sheet.title,
                "row_index": header_row_index,
                "start_column": get_column_letter(min_col),
                "end_column": get_column_letter(max_col),
                "cell_range": header_range,
                "region_type": "header",
                "headers": headers,
                "values": header_values,
                "text": (
                    f"[Sheet: {sheet.title} | Range: {header_range} | Row: {header_row_index} | Region: header]\n"
                    + "\n".join(header_pairs)
                ).strip(),
            })
            region_counts["header"] += 1

        for row_index in range(data_start_row, max_row + 1):
            values = _row_values(sheet, row_index, min_col, max_col)
            pairs = []
            for header, value in zip(headers, values):
                if value:
                    pairs.append(f"{header}: {value}")
            if not pairs:
                continue

            data_row_count += 1
            total_row_count += 1
            cell_range = f"{get_column_letter(min_col)}{row_index}:{get_column_letter(max_col)}{row_index}"
            region_type = _classify_region(headers, values, " | ".join(pairs))
            region_counts[region_type] += 1

            text_blocks.append(
                f"[Sheet: {sheet.title} | Range: {cell_range} | Row: {row_index} | Region: {region_type}]"
            )
            text_blocks.extend(pairs)
            text_blocks.append("")

            row_records.append({
                "sheet_name": sheet.title,
                "row_index": row_index,
                "start_column": get_column_letter(min_col),
                "end_column": get_column_letter(max_col),
                "cell_range": cell_range,
                "region_type": region_type,
                "headers": headers,
                "values": values,
                "text": (
                    f"[Sheet: {sheet.title} | Range: {cell_range} | Row: {row_index} | Region: {region_type}]\n"
                    + "\n".join(pairs)
                ).strip(),
            })

        sheets_summary.append({
            "sheet_name": sheet.title,
            "header_row_index": header_row_index,
            "headers": headers,
            "data_row_count": data_row_count,
            "region_counts": dict(region_counts),
        })

    if merged_present:
        warnings.append("Merged cells detected; only anchor-cell values were preserved.")
    warnings.extend(header_detection_warnings)

    aggregate_region_counts = dict(sum((Counter(sheet.get("region_counts", {})) for sheet in sheets_summary), Counter()))

    if total_row_count == 0:
        return {
            "success": False,
            "text": "",
            "file_type": file_type,
            "method": "openpyxl_structured",
            "error": "XLSX workbook contained no extractable rows.",
            "sheet_count": len(visible_sheets),
            "row_count": 0,
            "column_count_by_sheet": column_count_by_sheet,
            "sheets": sheets_summary,
            "sheet_names": sheet_names,
            "skipped_objects": ["charts", "slicers", "shapes", "images"],
            "warnings": warnings,
            "header_detection_used": any(sheet["header_row_index"] is not None for sheet in sheets_summary),
            "header_detection_warnings": header_detection_warnings,
            "row_records": row_records,
            "region_counts": aggregate_region_counts,
        }

    return {
        "success": True,
        "text": "\n".join(text_blocks).strip(),
        "file_type": file_type,
        "method": "openpyxl_structured",
        "error": None,
        "sheet_count": len(visible_sheets),
        "row_count": total_row_count,
        "column_count_by_sheet": column_count_by_sheet,
        "sheets": sheets_summary,
        "sheet_names": sheet_names,
        "skipped_objects": ["charts", "slicers", "shapes", "images"],
        "warnings": warnings,
        "header_detection_used": any(sheet["header_row_index"] is not None for sheet in sheets_summary),
        "header_detection_warnings": header_detection_warnings,
        "row_records": row_records,
        "region_counts": aggregate_region_counts,
    }
